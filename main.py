import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QDialog
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QLabel
from numpy import loadtxt
import numpy as np
import math
import serial
import serial.tools.list_ports
import struct
import matplotlib.pyplot as plt #pip install matplotlib
from scipy.misc import electrocardiogram #pip install scipy
from openpyxl import Workbook #pip install openpyxl
from openpyxl import load_workbook

def getecgplot_atrial(freq,title):
    ecg = electrocardiogram()
    frequency = freq
    time_data = np.arange(ecg.size) / frequency
    plt.title(title+"\tOrange: Atrial\tBlue: Ventricular")
    plt.plot(time_data, ecg)
    plt.xlabel("Time in seconds")
    plt.ylabel("Amplitude")
    plt.xlim(9, 10.2)
    plt.ylim(-1, 1.5)
    plt.show()


def getecgplot_ventricular(freq,title):
    ecg = electrocardiogram()
    frequency = freq
    time_data = np.arange(ecg.size) / (2*frequency)
    plt.title(title+"\tOrange: Atrial\tBlue: Ventricular")
    plt.plot(time_data, ecg)
    plt.xlabel("Time in seconds")
    plt.ylabel("Amplitude")
    plt.xlim(9, 10.2)
    plt.ylim(-1, 1.5)
    plt.show()

# this is the port being used when board is connected
frdm_port = "COM3"

class Mainscreen(QDialog):
    def __init__(self):
        super(Mainscreen,self).__init__()
        loadUi("mainscreen.ui",self)
        self.continuebutton.clicked.connect(self.gotologin)
        widget = QtWidgets.QStackedWidget()
        widget.setFixedWidth(500)
        widget.setFixedHeight(500)

    def gotologin(self):
        nextscreen = Login()
        widget.addWidget(nextscreen)
        widget.setCurrentIndex(widget.currentIndex()+1)

class Login(QDialog):
    def __init__(self):
        super(Login,self). __init__()
        loadUi("login.ui",self)
        self.password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.loginbutton.clicked.connect(self.loginfunction)
        widget.setFixedWidth(500)
        widget.setFixedHeight(500)
        self.createaccbutton.clicked.connect(self.gotocreate)
        self.invalid_error.setVisible(False)
        self.maxerror_2.setVisible(False)



    def loginfunction(self):
        db = open("Storage/database.txt", "r")

        global username
        global password

        username = self.username.text()
        password = self.password.text()



        if not len(username or password) < 1:
            user_store = []
            passw_store = []

            for i in db:
                try:
                    user,passw = i.split(",")
                    passw = passw.strip()
                    user_store.append(user)
                    passw_store.append(passw)
                    self.maxerror_2.setVisible(False)
                except:
                    self.maxerror_2.setVisible(True)



            data = dict(zip(user_store,passw_store))

            try:
                if data[username]:
                    try:
                        if password == data[username]:
                            print("Login sucsess")
                            print("Hi", username)
                            self.gotodash()
                        else:
                            print("Invalid credentials")
                            self.invalid_error.setVisible(True)

                    except:
                        print("Invalid credentials")
                        self.invalid_error.setVisible(True)

                else:
                    print("User does not exist")
                    self.invalid_error.setVisible(True)

            except:
                print("login error")
                self.invalid_error.setVisible(True)


    def gotocreate(self):
        createacc = CreateAcc()
        widget.addWidget(createacc)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def gotodash(self):
        dashboard = Dash()
        widget.addWidget(dashboard)
        widget.setCurrentIndex(widget.currentIndex()+1)

class CreateAcc(QDialog):
    def __init__(self):
        super(CreateAcc, self).__init__()
        loadUi("createacc.ui",self)
        self.submitbutton.clicked.connect(self.validateUser)
        #self.submitbutton.clicked.connect(self.createaccfunction)
        self.returnbutton.clicked.connect(self.returnfunction)
        self.password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.confirmpass.setEchoMode(QtWidgets.QLineEdit.Password)
        self.matcherror.setVisible(False)
        self.usererror.setVisible(False)
        self.blankerror.setVisible(False)
        self.maxerror.setVisible(False)
        self.maxerror_2.setVisible(False)
        self.openerror.setVisible(False)
        widget.setFixedWidth(500)
        widget.setFixedHeight(500)

    def returnfunction(self):
        login = Login()
        widget.addWidget(login)
        widget.setCurrentIndex(widget.currentIndex()+1)

    # this is to choose what characters are not valid to put. If comma is put then account wont be created.
    def validateUser(self):
        username = self.username.text()
        password = self.password.text()
        # comma is not valid
        invalid = [","]

        # in password and username if comma is entered
        if any(substring in username for substring in invalid) or any(substring in password for substring in invalid):
            print("Invalid username")
            print("Invalid password")
            self.maxerror_2.setVisible(True)

        else:
            db = open("Storage/database.txt")
            username = self.username.text()
            password = self.password.text()
            confirm_pass = self.confirmpass.text()

            if not len(username or password) < 1:
                user_store = []
                passw_store = []

                try:
                    for i in db:
                        user,passw = i.split(",")
                        passw = passw.strip()
                        user_store.append(user)
                        passw_store.append(passw)


                except:
                    self.maxerror_2.setVisible(True)


                data = dict(zip(user_store,passw_store))

                if password != confirm_pass:
                    print("Passwords do not match")
                    self.confirmpass.clear()
                    self.matcherror.setVisible(True)
                    return

                elif username in user_store:
                    print("User already exists, choose another")
                    self.username.clear()
                    self.usererror.setVisible(True)
                    return

                elif username == "" or password =="" or confirm_pass == "":
                    print("Cannot leave blank fields")
                    self.username.clear()
                    self.password.clear()
                    self.confirmpass.clear()
                    self.blankerror.setVisible(True)
                    return


                else:
                    db = open("Storage/database.txt", "r")
                    read_db = db.readlines()
                    db.close()

                    if len(read_db) < 10:

                        try:
                            wb = load_workbook("AccountData.xlsx")
                            wb.copy_worksheet(wb["Template"]).title = username
                            wb.save("AccountData.xlsx")

                            db = open("Storage/database.txt", "a")
                            db.write(username+", "+ password+"\n")
                            print("Successfully created account:", username," And added to excel database with title:", username)
                            db.close()

                        except:
                            #SHOW FATAL ERROR MESSAGE
                            self.username.clear()
                            self.password.clear()
                            self.confirmpass.clear()
                            print("FATAL ERROR: CLOSE EXCEL FILE")
                            self.openerror.setVisible(True)
                            return

                    else:
                        print("User Limit Reached")
                        self.username.clear()
                        self.password.clear()
                        self.confirmpass.clear()
                        self.maxerror.setVisible(True)
                        return


                login=Login()
                widget.addWidget(login)
                widget.setCurrentIndex(widget.currentIndex()+1)

class Dash(QDialog):
    def __init__(self):
        super(Dash, self).__init__()
        loadUi("dashboard.ui", self)
        widget.setFixedWidth(1200)
        widget.setFixedHeight(600)
        self.logoutbutton.clicked.connect(self.logoutfunction)
        self.VOObutton.clicked.connect(self.gotovoo)
        self.AOObutton.clicked.connect(self.gotoaoo)
        self.AAIbutton.clicked.connect(self.gotoaai)
        self.VVIbutton.clicked.connect(self.gotovvi)
        self.AOORbutton.clicked.connect(self.gotoaoor)
        self.VOORbutton.clicked.connect(self.gotovoor)
        self.AAIRbutton.clicked.connect(self.gotoaair)
        self.VVIRbutton.clicked.connect(self.gotovvir)
        self.ECGPLOTbutton.clicked.connect(self.gotoecgplot)
        self.refreshbutton.clicked.connect(self.deviceconnected)
        self.refreshbutton.clicked.connect(self.deviceverification)
        self.disconnectedmsg.setVisible(True)
        self.connectedmsg.setVisible(False)

    # check if board is connected
    def deviceconnected(self):
        connect_true = serial.tools.list_ports.comports()
       # check if plugged into correct board
        try:
            for w in connect_true:
                if(w.device == "COM3"):
                    frdm_port = w.device
                    con = True
                    print("connected")
                    self.connectedmsg.setVisible(True)
                    self.disconnectedmsg.setVisible(False)

            return [frdm_port,con]

        except:

            print("notconnected")
            self.disconnectedmsg.setVisible(True)
            self.connectedmsg.setVisible(False)
            return [None, False]

    def deviceverification(self):


        for port in serial.tools.list_ports.comports():
            db = open("Storage/Verification.txt")
            manufacturer = port.manufacturer
            serialnumber = port.serial_number
            print(manufacturer, serialnumber)
            man_store = []
            ser_store = []

            if manufacturer is not None and serialnumber is not None:
                for i in db:
                    man,ser = i.split(",")
                    ser = ser.strip()
                    man_store.append(man)
                    ser_store.append(ser)


                data = dict(zip(man_store,ser_store))

                if serialnumber in ser_store:
                    print("Serial Already Logged In")
                    #display current serial
                    return

                else:
                    db = open("Storage/Verification.txt", "r")
                    read_db = db.readlines()
                    db.close()
                    if manufacturer is not None and serialnumber is not None:
                        db = open("Storage/Verification.txt", "a")
                        db.write(manufacturer+", "+ serialnumber+"\n")
                        print("Successfully Stored Serial Number")
                        #add this to login screen
                        db.close()

            f = open('Storage/Verification.txt', 'r')
            content = f.read()
            print(content)
            f.close()

    def gotoecgplot(self):
        ecg = electrocardiogram()
        frequency = 360
        time_data = np.arange(ecg.size) / frequency
        plt.title("Electrocardiogram")
        plt.plot(time_data, ecg)
        plt.xlabel("Time in seconds")
        plt.ylabel("ECG in milli Volts")
        plt.xlim(9, 10.2)
        plt.ylim(-1, 1.5)
        plt.show()

    def gotovoo(self):
        voo = VOO()
        widget.addWidget(voo)
        widget.setCurrentIndex(widget.currentIndex()+1)


        VOOUP = voo.UPLIMIT.text()

        try:
            wb = load_workbook("AccountData.xlsx")
            ws = wb[username]

            avoo = ws["D2"]
            voo.LRL.setText(str(avoo.value))

            bvoo = ws["D3"]
            voo.UPLIMIT.setText(str(bvoo.value))

            cvoo = ws["D5"]
            voo.VPW.setText(str(cvoo.value))

            dvoo = ws["D4"]
            voo.VA.setText(str(dvoo.value))  
        except:
            print("could not add number")
            voo.UPLIMIT.setText(str(0))
            voo.LRL.setText(str(0))
            voo.VPW.setText(str(0))
            voo.VA.setText(str(0))       
                               

    def gotoaoo(self):
        aoo = AOO()
        widget.addWidget(aoo)
        widget.setCurrentIndex(widget.currentIndex()+1)

        # file is opened and the values are sent back to the parameters
        # values are saved unique to user 
        
        try:

            wb = load_workbook("AccountData.xlsx")
            ws = wb[username]

            aAOO = ws["B2"]
            aoo.AOOLRL.setText(str(aAOO.value))

            bAOO = ws["B3"]
            aoo.AOOUP.setText(str(bAOO.value))

            cAOO = ws["B7"]
            aoo.AOOPW.setText(str(cAOO.value))

            dAOO = ws["B6"]
            aoo.AOOAA.setText(str(dAOO.value))


        except:
            print("could not add number")
            aoo.AOOUP.setText(str(0))
            aoo.AOOLRL.setText(str(0))
            aoo.AOOPW.setText(str(0))
            aoo.AOOAA.setText(str(0))



    def gotoaai(self):
        aai = AAI()
        widget.addWidget(aai)
        widget.setCurrentIndex(widget.currentIndex()+1)


        wb = load_workbook("AccountData.xlsx")
 
        try:
            wb = load_workbook("AccountData.xlsx")
         
            ws = wb[username]
            Aaai = ws['C2']
            aai.AAILRL.setText(str(Aaai.value))
            
            Baai =ws['C3'] 
            aai.AAIURL.setText(str(Baai.value))
           
            Caai = ws['C6'] 
            aai.AAIAA.setText(str(Caai.value))
          
            Daai = ws['C7']
            aai.AAIAPW.setText(str(Daai.value))
        
            Eaai = ws['C8'] 
            aai.AAIAS.setText(str(Eaai.value))
         
            Faai = ws['C9'] 
            aai.AAIARP.setText(str(Faai.value))
          
            Gaai = ws['C10'] 
            aai.AAIPVARP.setText(str(Gaai.value))
            
            Haai = ws['C11'] 
            aai.AIIH.setText(str(Haai.value))
          
            Iaai = ws['C19'] 
            aai.AAIRS.setText(str(Iaai.value))


        except:
            #if nothing is in the textfile give the value of 0
            aai.AAILRL.setText(str(0))
            aai.AAIURL.setText(str(0))
            aai.AAIAA.setText(str(0))
            aai.AAIAPW.setText(str(0))
            aai.AAIAS.setText(str(0))
            aai.AAIARP.setText(str(0))
            aai.AAIPVARP.setText(str(0))
            aai.AIIH.setText(str(0))
            aai.AAIRS.setText(str(0))


    def gotovvi(self):
        vvi = VVI()
        widget.addWidget(vvi)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
        wb = load_workbook("AccountData.xlsx")
        # open the file and sent the numbers inputted back into the mode
        #values are sent to the correct parameters
        try:
            wb = load_workbook("AccountData.xlsx")
         
            ws = wb[username]
            Avvi = ws['E2']
            vvi.VVILRL.setText(str(Avvi.value))

            Bvvi =ws['E3'] 
            vvi.VVIURL.setText(str(Bvvi.value))
            
            Cvvi = ws['E6'] 
            vvi.VVIVA.setText(str(Cvvi.value))

            Dvvi = ws['E7']
            vvi.VVIVPW.setText(str(Dvvi.value))

            Evvi = ws['E8'] 
            vvi.VVIVS.setText(str(Evvi.value))

            Fvvi = ws['E13'] 
            vvi.VVIVRP.setText(str(Fvvi.value))
            
            Gvvi = ws['E11'] 
            vvi.VVIH.setText(str(Gvvi.value))
            
            Hvvi = ws['E19'] 
            vvi.VVIRS.setText(str(Hvvi.value))




        except:
            #if nothing is in the textfile give the value of 0
            vvi.VVILRL.setText(str(0))
            vvi.VVIURL.setText(str(0))
            vvi.VVIVA.setText(str(0))
            vvi.VVIVPW.setText(str(0))
            vvi.VVIVS.setText(str(0))
            vvi.VVIVRP.setText(str(0))
            vvi.VVIH.setText(str(0))
            vvi.VVIRS.setText(str(0))


    def gotoaoor(self):
        aoor = AOOR()
        widget.addWidget(aoor)
        widget.setCurrentIndex(widget.currentIndex()+1)


        # open the textfile and send the values to array
        # when user reclicks mode without logging out values stay
        with open('Storage/AOOR.txt', 'r') as file:
            datasAOOR = file.read()
        datasAOORvalues = datasAOOR.split("\n")

        #values are sent to the correct parameters
        try:
            aAOOR = datasAOORvalues[0]

            aoor.AOORLRL.setText(str(aAOOR))

            bAOOR = datasAOORvalues[1]
            aoor.AOORURL.setText(str(bAOOR))

            gAOOR = datasAOORvalues[2]
            aoor.AOORAA.setText(str(gAOOR))

            cAOOR = datasAOORvalues[3]
            aoor.AOORPW.setText(str(cAOOR))

            dAOOR = datasAOORvalues[4]
            aoor.AOORMSR.setText(str(dAOOR))

            eAOOR = datasAOORvalues[5]
            aoor.ACTIVEAOOR.setText(str(eAOOR))

            fAOOR = datasAOORvalues[6]
            aoor.ReactTimeAOOR.setText(str(fAOOR))

            hAOOR = datasAOORvalues[7]
            aoor.RespFactAOOR.setText(str(hAOOR))

            iAOOR = datasAOORvalues[8]
            aoor.RecovTimeAOOR.setText(str(iAOOR))



        #if nothing is in the textfile give the value of 0
        except:
            aoor.AOORLRL.setText(str(0))
            aoor.AOORURL.setText(str(0))
            aoor.AOORPW.setText(str(0))
            aoor.AOORMSR.setText(str(0))
            aoor.ACTIVEAOOR.setText(str(0))
            aoor.ReactTimeAOOR.setText(str(0))
            aoor.AOORAA.setText(str(0))
            aoor.RespFactAOOR.setText(str(0))
            aoor.RecovTimeAOOR.setText(str(0))


    def gotovoor(self):
        voor = VOOR()
        widget.addWidget(voor)
        widget.setCurrentIndex(widget.currentIndex()+1)

        # open the textfile and send the values to array
        # when user reclicks mode without logging out values stay
        with open('Storage/VOOR.txt', 'r') as file:
            datasVOOR = file.read()
        datasVOORvalues = datasVOOR.split("\n")

        #values are sent to the correct parameters
        try:
            aVOOR = datasVOORvalues[0]
            voor.VOORLRL.setText(str(aVOOR))

            bVOOR = datasVOORvalues[1]
            voor.VOORURL.setText(str(bVOOR))


            cVOOR = datasVOORvalues[3]
            voor.VOORPW.setText(str(cVOOR))

            dVOOR = datasVOORvalues[4]
            voor.VOORMSR.setText(str(dVOOR))

            eVOOR = datasVOORvalues[5]
            voor.ACTIVEVOOR.setText(str(eVOOR))

            fVOOR = datasVOORvalues[6]
            voor.ReactTimeVOOR.setText(str(fVOOR))


            gVOOR = datasVOORvalues[2]
            voor.VOORVA.setText(str(gVOOR))

            hVOOR = datasVOORvalues[7]
            voor.RespFactVOOR.setText(str(hVOOR))

            iVOOR = datasVOORvalues[8]
            voor.RecovTimeVOOR.setText(str(iVOOR))



        #if nothing is in the textfile give the value of 0
        except:
            voor.VOORLRL.setText(str(0))
            voor.VOORURL.setText(str(0))
            voor.VOORPW.setText(str(0))
            voor.VOORMSR.setText(str(0))
            voor.ACTIVEVOOR.setText(str(0))
            voor.ReactTimeVOOR.setText(str(0))
            voor.VOORVA.setText(str(0))
            voor.RespFactVOOR.setText(str(0))
            voor.RecovTimeVOOR.setText(str(0))

    def gotoaair(self):
        aair = AAIR()
        widget.addWidget(aair)
        widget.setCurrentIndex(widget.currentIndex()+1)

        # open the textfile and send all the values to the array
        with open('Storage/AAIR.txt', 'r') as file:
            datasAAIR = file.read()
        datasAAIRvalues = datasAAIR.split("\n")

        # send all the values to the respective parameters
        try:
            aAAIR = datasAAIRvalues[0]
            aair.AAIRLRL.setText(str(aAAIR))

            bAAIR = datasAAIRvalues[1]
            aair.AAIRURL.setText(str(bAAIR))

            cAAIR = datasAAIRvalues[2]
            aair.AAIRAA.setText(str(cAAIR))

            dAAIR = datasAAIRvalues[4]
            aair.AAIRPW.setText(str(dAAIR))

            eAAIR = datasAAIRvalues[3]
            aair.AAIRMSR.setText(str(eAAIR))

            fAAIR = datasAAIRvalues[5]
            aair.AAIRAS.setText(str(fAAIR))

            gAAIR = datasAAIRvalues[6]
            aair.AAIRARP.setText(str(gAAIR))

            hAAIR = datasAAIRvalues[7]
            aair.ReactTimeAAIR.setText(str(hAAIR))

            iAAIR = datasAAIRvalues[8]
            aair.AAIRPVARP.setText(str(iAAIR))

            jAAIR = datasAAIRvalues[9]
            aair.AAIRHY.setText(str(jAAIR))

            kAAIR = datasAAIRvalues[10]
            aair.AAIRRS.setText(str(kAAIR))

            lAAIR = datasAAIRvalues[11]
            aair.RespFactAAIR.setText(str(lAAIR))

            mAAIR = datasAAIRvalues[12]
            aair.ACTIVEAAIR.setText(str(mAAIR))

            nVOOR = datasAAIRvalues[13]
            aair.RecovTimeAAIR.setText(str(nVOOR))



        #if nothing is in the textfile give the value of 0
        except:
            aair.AAIRLRL.setText(str(0))
            aair.AAIRURL.setText(str(0))
            aair.AAIRAA.setText(str(0))
            aair. AAIRPW.setText(str(0))
            aair.AAIRMSR.setText(str(0))
            aair.AAIRAS.setText(str(0))
            aair.AAIRARP.setText(str(0))
            aair.ReactTimeAAIR.setText(str(0))
            aair.AAIRPVARP.setText(str(0))
            aair.AAIRHY.setText(str(0))
            aair.AAIRRS.setText(str(0))
            aair.RespFactAAIR.setText(str(0))
            aair.ACTIVEAAIR.setText(str(0))
            aair.RecovTimeAAIR.setText(str(0))

    def gotovvir(self):
        vvir = VVIR()
        widget.addWidget(vvir)
        widget.setCurrentIndex(widget.currentIndex()+1)

        # open the textfile and send all the values to the array
        # when user reclicks mode without logging out values stay
        with open('Storage/VVIR.txt', 'r') as file:
            datasVVIR = file.read()
        datasVVIRvalues = datasVVIR.split("\n")

         # send all the values to the respective parameters
        try:
            aVVIR = datasVVIRvalues[0]
            vvir.VVIRLRL.setText(str(aVVIR))

            bVVIR = datasVVIRvalues[1]
            vvir.VVIRURL.setText(str(bVVIR))

            cVVIR = datasVVIRvalues[2]
            vvir.VVIRVA.setText(str(cVVIR))

            dVVIR = datasVVIRvalues[3]
            vvir.VVIRPW.setText(str(dVVIR))

            eVVIR = datasVVIRvalues[4]
            vvir.VVIRMSR.setText(str(eVVIR))

            fVVIR = datasVVIRvalues[5]
            vvir.VVIRVS.setText(str(fVVIR))

            gVVIR = datasVVIRvalues[6]
            vvir.VVIRIVRP.setText(str(gVVIR))

            # hVVIR = datasVVIRvalues[7]
            # vvir.VVIRVS.setText(str(hVVIR))

            iVVIR = datasVVIRvalues[7]
            vvir.VVIRReactTime.setText(str(iVVIR))

            jVVIR = datasVVIRvalues[8]
            vvir.VVIRHY.setText(str(jVVIR))

            kVVIR = datasVVIRvalues[9]
            vvir.VVIRRS.setText(str(kVVIR))

            lAAIR = datasVVIRvalues[10]
            vvir.RespFactVVIR.setText(str(lAAIR))

            mAAIR = datasVVIRvalues[11]
            vvir.ACTIVEVVIR.setText(str(mAAIR))

            nAAIR = datasVVIRvalues[12]
            vvir.RecovTimeVVIR.setText(str(nAAIR))


        #if nothing is in the textfile give the value of 0
        except:

            vvir.VVIRLRL.setText(str(0))
            vvir.VVIRURL.setText(str(0))
            vvir.VVIRVA.setText(str(0))
            vvir.VVIRPW.setText(str(0))
            vvir.VVIRMSR.setText(str(0))
            vvir.VVIRVS.setText(str(0))
            vvir.VVIRIVRP.setText(str(0))
            vvir.VVIRReactTime.setText(str(0))
            vvir.VVIRHY.setText(str(0))
            vvir.VVIRRS.setText(str(0))
            vvir.RespFactVVIR.setText(str(0))
            vvir.ACTIVEVVIR.setText(str(0))
            vvir.RecovTimeVVIR.setText(str(0))

    def logoutfunction(self):
        logout = Mainscreen()
        widget.setFixedWidth(500)
        widget.setFixedHeight(500)
        widget.addWidget(logout)
        print ("Account Logged Out") #self.message.setVisible(True)

        widget.setCurrentIndex(widget.currentIndex()+1)


class VOO(QDialog):
    def __init__(self):
        super(VOO, self).__init__()
        loadUi("VOO.ui", self)
        self.submitbutton.clicked.connect(self.inputfunction)
        self.transferdatabutton.clicked.connect(self.transferfunction)
        self.backbutton.clicked.connect(self.backfunction)


        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        widget.setFixedWidth(900)
        widget.setFixedHeight(600)



    #function to handel the ranges for modes
    def inputfunction(self):

        global VOOLRL
        global VOOUP
        global VA
        global VOOVPW

        VOOLRL = self.LRL.text()
        VOOUP = self.UPLIMIT.text()
        VOOVPW = self.VPW.text()
        VA = self.VA.text()



        try:
            # ranges for the voo. If range is not met then show invalid error
            if ((((int(VOOLRL)) >= 30) and ((int(VOOLRL)) <= 49) and (int(VOOLRL) % 5 == 0))) or ((((int(VOOLRL)) >= 50) and ((int(VOOLRL)) <= 89) and (int(VOOLRL) % 1 == 0))) or ((((int(VOOLRL)) >= 90) and ((int(VOOLRL)) <= 175) and (int(VOOLRL) % 5 == 0))):
                if(((int(VOOUP)) >= 50) and (int(VOOUP)) <=175) and (int(VOOUP) % 5 == 0):
                    if(((((float(VA)) >= 0.1) and (float(VA)) <=5) and (10*(float(VA)) % 1 == 0)) or (((float(VA)) == 0))):
                        if((((int(VOOVPW)) >= 1)) and ((int(VOOVPW)) <= 30) and (10*(float(VOOVPW)) % 10 == 0)):
                            # open to the file and write the inputed numbers
                            self.INVALID.setVisible(False)
                            self.SUCCESS.setVisible(True)
                            getecgplot_atrial(360,"VOO Plot")
                            getecgplot_ventricular(360,"VOO Plot")
                            wb = load_workbook("AccountData.xlsx")

                            try:
                                ws = wb[username]
                                ws['D2'] = VOOLRL
                                ws['D3'] = VOOUP
                                ws['D5'] = VOOVPW
                                ws['D4'] = VA

                                wb.save("AccountData.xlsx")
                                print("Added number VOO Data")
                                #SHOW THE TRANSFER DATA BUTON AFTER THIS
                                


                            except:
                                print("Couldn't add the number")

                        else:
                          self.INVALID.setVisible(True)
                          self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)



        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)

#         global VOOLRL
#         global VOOUP
#         global VA
#         global VOOVPW

    def transferfunction(self):
        #SET VOO VALUES AND PUSH THEM TO BOARD
        # try:
        frdm_port = "COM3"

        wb = load_workbook("AccountData.xlsx")
        ws = wb[username]
        VOOVPW = ws['D5']
        VOOLRL = ws['D2']
        VA = ws['D4']

        Start = b'\x16' #putting a value in bits into a variable bcuz the var will be added to the byte stream and sent over to simulink
        SYNC = b'\x22'
        #Fn_set = b'\x55'
        Fn_set = struct.pack("B",55)
        p_aPaceWidth = struct.pack('B', 7)
        p_vPaceWidth = struct.pack('B', int(VOOVPW.value))
        p_aPaceAmp = struct.pack('f', 1.2)#f is single
        p_vPaceAmp = struct.pack('f', float(VA.value))
        p_atrialsensitivity = struct.pack('f', 4)
        p_ventriclesensitivity = struct.pack('f', 9)
        p_ARP = struct.pack('H', 200) #H is uint16
        p_VRP = struct.pack('H', 300)
        p_lowratelimit = struct.pack('B', int(VOOLRL.value)) #we used B becasue its uint8
        p_Mode = struct.pack('B', 2)
        #blue_en = struct.pack("B", 1)

        Signal_set = Start + Fn_set + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode
        Signal_echo = Start + SYNC + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode

        with serial.Serial(frdm_port, 115200) as pacemaker: #pushes signal to the baord
            pacemaker.write(Signal_set)

        with serial.Serial(frdm_port, 115200) as pacemaker: #send the byte stream that echos to the pacemaker (the block that sends parameter back to you)
            pacemaker.write(Signal_echo)
            data = pacemaker.read(24)
            p_aPaceWidth = data[0]
            p_vPaceWidth = data[1]
            p_aPaceAmp = struct.unpack('f', data[2:6])[0] 
            p_vPaceAmp = struct.unpack('f', data[6:10])[0]
            p_atrialsensitivity = struct.unpack('f', data[10:14])[0]
            p_ventriclesensitivity = struct.unpack('f', data[14:18])[0]
            p_ARP = struct.unpack('H', data[18:20])[0]  #H is uint16
            p_VRP = struct.unpack('H', data[20:22])[0] 
            p_lowratelimit = data[22] #we used B becasue its uint8
            p_Mode = data[23]
            

        print("From the board:")
        print("p_aPaceWidth = ", p_aPaceWidth)
        print("p_vPaceWidth = ", p_vPaceWidth)
        print("p_aPaceAmp = ", p_aPaceAmp)
        print("p_vPaceAmp = ",  p_vPaceAmp)
        print("p_atrialsensitivity = ", p_atrialsensitivity)
        print("p_ventriclesensitivity = ",  p_ventriclesensitivity)
        print("p_ARP = ",  p_ARP)
        print("p_VRP = ",  p_VRP)
        print("p_lowratelimit = ",  p_lowratelimit)
        print("p_Mode = ",  p_Mode)

        
    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class AOO(QDialog):
    def __init__(self):
        super(AOO, self).__init__()
        loadUi("AOO.ui", self)
        widget.setFixedWidth(900)
        widget.setFixedHeight(600)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.AOOsubmitbutton.clicked.connect(self.AOOinputfunction)
        self.transferdatabutton.clicked.connect(self.transferfunction)
        self.backbutton.clicked.connect(self.backfunction)

    def AOOinputfunction (self):

        global AOOLRL
        global AOOUP
        global AOOPW
        global AOOAA


        AOOLRL = self.AOOLRL.text()
        AOOUP = self.AOOUP.text()
        AOOPW = self.AOOPW.text()
        AOOAA = self.AOOAA.text()
        try:
            # ranges for the voo. If range is not met then show invalid error
            if ((((int(AOOLRL)) >= 30) and ((int(AOOLRL)) <= 49) and (int(AOOLRL) % 5 == 0))) or (((int((AOOLRL)) >= 50) and ((int(AOOLRL)) <= 90) and (int(AOOLRL) % 1 == 0))) or ((((int(AOOLRL)) >= 91) and ((int(AOOLRL)) <= 175) and (int(AOOLRL) % 5 == 0))):
                if(((int(AOOUP)) >= 50) and (int(AOOUP)) <=175) and (int(AOOUP) % 5 == 0):
                    if((((float(AOOAA)) >= 0.1) and ((float(AOOAA)) <=5) and (10*(float(AOOAA)) % 1 == 0))  or (float(AOOAA)) == 0):
                        if ((((float(AOOPW)) >= 1)) and ((float(AOOPW)) <= 30) and ((float(AOOPW)) % 1 == 0)):
                            self.INVALID.setVisible(False)
                            self.SUCCESS.setVisible(True)
                            # open to the file and write the inputed numbers
                            getecgplot_atrial(720,"AOO Plot")
                            getecgplot_ventricular(720,"AOO Plot")
                            wb = load_workbook("AccountData.xlsx")

                            try:
                                ws = wb[username]
                                ws['B2'] = AOOLRL
                                ws['B3'] = AOOUP
                                ws['B6'] = AOOAA
                                ws['B7'] = AOOPW

                                wb.save("AccountData.xlsx")
                                print("Added number AOO Data")


                            except:
                                print("Couldn't add the number")

                        else:
                            self.INVALID.setVisible(True)
                            self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)

    
    def transferfunction(self):
        #SET AOO VALUES AND PUSH THEM TO BOARD
        # try:
        frdm_port = "COM3"

        wb = load_workbook("AccountData.xlsx")
        ws = wb[username]
        AOOLRL = ws["B2"]
        AOOUP = ws["B3"]
        AOOPW = ws["B7"]
        AOOAA = ws["B6"]

        Start = b'\x16' #putting a value in bits into a variable bcuz the var will be added to the byte stream and sent over to simulink
        SYNC = b'\x22'
        #Fn_set = b'\x55'
        Fn_set = struct.pack("B",55)
        p_aPaceWidth = struct.pack('B', int(AOOPW.value))
        p_vPaceWidth = struct.pack('B', int(VOOVPW.value))
        p_aPaceAmp = struct.pack('f', float(AOOAA.value))#f is single
        p_vPaceAmp = struct.pack('f', float(VA.value))
        p_atrialsensitivity = struct.pack('f', 4)
        p_ventriclesensitivity = struct.pack('f', 9)
        p_ARP = struct.pack('H', 200) #H is uint16
        p_VRP = struct.pack('H', 300)
        p_lowratelimit = struct.pack('B', int(AOOLRL.value)) #we used B becasue its uint8
        p_Mode = struct.pack('B', 1)
        #blue_en = struct.pack("B", 1)

        Signal_set = Start + Fn_set + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode
        Signal_echo = Start + SYNC + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode

        with serial.Serial(frdm_port, 115200) as pacemaker: #pushes signal to the baord
            pacemaker.write(Signal_set)

        with serial.Serial(frdm_port, 115200) as pacemaker: #send the byte stream that echos to the pacemaker (the block that sends parameter back to you)
            pacemaker.write(Signal_echo)
            data = pacemaker.read(24)
            p_aPaceWidth = data[0]
            p_vPaceWidth = data[1]
            p_aPaceAmp = struct.unpack('f', data[2:6])[0] 
            p_vPaceAmp = struct.unpack('f', data[6:10])[0]
            p_atrialsensitivity = struct.unpack('f', data[10:14])[0]
            p_ventriclesensitivity = struct.unpack('f', data[14:18])[0]
            p_ARP = struct.unpack('H', data[18:20])[0]  #H is uint16
            p_VRP = struct.unpack('H', data[20:22])[0] 
            p_lowratelimit = data[22] #we used B becasue its uint8
            p_Mode = data[23]
            

        print("From the board:")
        print("p_aPaceWidth = ", p_aPaceWidth)
        print("p_vPaceWidth = ", p_vPaceWidth)
        print("p_aPaceAmp = ", p_aPaceAmp)
        print("p_vPaceAmp = ",  p_vPaceAmp)
        print("p_atrialsensitivity = ", p_atrialsensitivity)
        print("p_ventriclesensitivity = ",  p_ventriclesensitivity)
        print("p_ARP = ",  p_ARP)
        print("p_VRP = ",  p_VRP)
        print("p_lowratelimit = ",  p_lowratelimit)
        print("p_Mode = ",  p_Mode)
    
    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class AAI(QDialog):
    def __init__(self):
        super(AAI, self).__init__()
        loadUi("AAI.ui", self)
        self.AAIsubmit.clicked.connect(self.AAIinputfunction)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.backbutton.clicked.connect(self.backfunction)
        widget.setFixedWidth(900)
        widget.setFixedHeight(670)

    def AAIinputfunction (self):

        global AAILRL
        global AAIURL
        global AAIAW
        global AAIAPW
        global AAIAS
        global AAIARP
        global AAIPVARP
        global AIIH
        global AAIRS

        AAILRL = self.AAILRL.text()
        AAIURL = self.AAIURL.text()
        AAIAW = self.AAIAA.text()
        AAIAPW = self.AAIAPW.text()
        AAIAS = self.AAIAS.text()
        AAIARP = self.AAIARP.text()
        AAIPVARP = self.AAIPVARP.text()
        AIIH = self.AIIH.text()
        AAIRS = self.AAIRS.text()
        try:
            # ranges for the AAI. If range is not met then show invalid error else show sucess message
            if ((((int(AAILRL)) >= 30) and ((int(AAILRL)) <= 49) and (int(AAILRL) % 5 == 0))) or ((((int(AAILRL)) >= 50) and ((int(AAILRL)) <= 90) and (int(AAILRL) % 1 == 0))) or ((((int(AAILRL)) >= 91) and ((int(AAILRL)) <= 175) and (int(AAILRL) % 5 == 0))):
                if(((int(AAIURL)) >= 50) and (int(AAIURL)) <=175) and (int(AAIURL) % 5 == 0):
                    if(((((float(AAIAW)) >= 0.1) and (float(AAIAW)) <=5.0) and (10*(float(AAIAW)) % 1 == 0)) or ((int(AAIAW)) == 0)):
                        if(((((int(AAIAPW)) >= 1) and (int(AAIAPW)) <=30)) and ((int(AAIAPW)) %1 == 0)):
                            if(((float(AAIAS) >= 0) and (float(AAIAS) <= 5) and (10*float(AAIAS) % 1 == 0))):
                                if((int(AAIARP) >= 150) and (int(AAIARP) <= 500) and ((int(AAIARP)) % 10 == 0)):
                                    if( ((int(AAIPVARP) >= 150) and ((int(AAIPVARP) <= 500) and ((int(AAIPVARP)) % 10 == 0)))):
                                        if (((((int(AIIH)) >= 30) and ((int(AIIH)) <= 49) and (int( AIIH) % 5 == 0))) or ((((int( AIIH)) >= 50) and ((int(AIIH)) <= 89) and (int( AIIH) % 1 == 0))) or ((((int( AIIH)) >= 90) and ((int(AIIH)) <= 175) and (int(AIIH) % 5 == 0))) or ((int(AIIH) == 0))):
                                            if((((int(AAIRS) >= 0) and ((int(AAIRS) <= 21) and ((int(AAIRS)) % 3 == 0))))):
                                                self.INVALID.setVisible(False)
                                                self.SUCCESS.setVisible(True)
                                                getecgplot_atrial(640,"AAI Plot")
                                                getecgplot_ventricular(640,"AAI Plot")


                                                wb = load_workbook("AccountData.xlsx")

                                                try:
                                                    ws = wb[username]
                                                    ws['C2'] = AAILRL
                                                    ws['C3'] = AAIURL
                                                    ws['C6'] = AAIAW
                                                    ws['C7'] = AAIAPW
                                                    ws['C8'] = AAIAS
                                                    ws['C9'] = AAIARP
                                                    ws['C10'] = AAIPVARP
                                                    ws['C11'] =AIIH
                                                    ws['C19'] =AAIRS


                                                    wb.save("AccountData.xlsx")
                                                    print("Added AAI Data")


                                                except:
                                                    print("Couldn't add the number")


                                            else:
                                                self.INVALID.setVisible(True)
                                                self.SUCCESS.setVisible(False)

                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)

                                else:

                                    # self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                        else:
                             self.INVALID.setVisible(True)
                             self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)

    def transferfunction(self):
        #SET AAI VALUES AND PUSH THEM TO BOARD
        # try:
        frdm_port = "COM3"

        wb = load_workbook("AccountData.xlsx")
        ws = wb[username]

        AAILRL = ws['C2']
        AAIURL =  ws['C3'] 
        AAIAW = ws['C6']  
        AAIAPW =  ws['C7'] 
        AAIAS = ws['C8']  
        AAIARP = ws['C9'] 
        AAIPVARP = ws['C10']  
        AIIH = ws['C11'] 
        AAIRS = ws['C19']

        Start = b'\x16' #putting a value in bits into a variable bcuz the var will be added to the byte stream and sent over to simulink
        SYNC = b'\x22'
        #Fn_set = b'\x55'
        Fn_set = struct.pack("B",55)
        p_aPaceWidth = struct.pack('B', int(AAIAW.value))
        p_vPaceWidth = struct.pack('B', int(VOOVPW.value))
        p_aPaceAmp = struct.pack('f', float(AAIAW.value))#f is single
        p_vPaceAmp = struct.pack('f', float(AAIAPW.value))
        p_atrialsensitivity = struct.pack('f', float(AAIAS.value))
        p_ventriclesensitivity = struct.pack('f', 9)
        p_ARP = struct.pack('H', int(AAIARP.value)) #H is uint16
        p_VRP = struct.pack('H', int(AAIPVARP.value))
        p_lowratelimit = struct.pack('B', int(AAILRL.value)) #we used B becasue its uint8
        p_Mode = struct.pack('B', 3)
        #blue_en = struct.pack("B", 1)

        Signal_set = Start + Fn_set + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode
        Signal_echo = Start + SYNC + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode

        with serial.Serial(frdm_port, 115200) as pacemaker: #pushes signal to the baord
            pacemaker.write(Signal_set)

        with serial.Serial(frdm_port, 115200) as pacemaker: #send the byte stream that echos to the pacemaker (the block that sends parameter back to you)
            pacemaker.write(Signal_echo)
            data = pacemaker.read(24)
            p_aPaceWidth = data[0]
            p_vPaceWidth = data[1]
            p_aPaceAmp = struct.unpack('f', data[2:6])[0] 
            p_vPaceAmp = struct.unpack('f', data[6:10])[0]
            p_atrialsensitivity = struct.unpack('f', data[10:14])[0]
            p_ventriclesensitivity = struct.unpack('f', data[14:18])[0]
            p_ARP = struct.unpack('H', data[18:20])[0]  #H is uint16
            p_VRP = struct.unpack('H', data[20:22])[0] 
            p_lowratelimit = data[22] #we used B becasue its uint8
            p_Mode = data[23]
            

        print("From the board:")
        print("p_aPaceWidth = ", p_aPaceWidth)
        print("p_vPaceWidth = ", p_vPaceWidth)
        print("p_aPaceAmp = ", p_aPaceAmp)
        print("p_vPaceAmp = ",  p_vPaceAmp)
        print("p_atrialsensitivity = ", p_atrialsensitivity)
        print("p_ventriclesensitivity = ",  p_ventriclesensitivity)
        print("p_ARP = ",  p_ARP)
        print("p_VRP = ",  p_VRP)
        print("p_lowratelimit = ",  p_lowratelimit)
        print("p_Mode = ",  p_Mode)    
    
    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class VVI(QDialog):
    def __init__(self):
        super(VVI, self).__init__()
        loadUi("VVI.ui", self)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.VVIsubmitbutton.clicked.connect(self.VVIinputfunction)
        self.backbutton.clicked.connect(self.backfunction)
        self.transferdatabutton.clicked.connect(self.transferfunction)
        widget.setFixedWidth(900)
        widget.setFixedHeight(700)

    def VVIinputfunction(self):

        global VVILRL
        global VVIURL
        global VVIVA
        global VVIVPW
        global VVIVS
        global VVIVRP
        global VVIH 
        global VVIRS 
        
        VVILRL = self.VVILRL.text()
        VVIURL = self.VVIURL.text()
        VVIVA = self.VVIVA.text()
        VVIVPW = self.VVIVPW.text()
        VVIVS = self.VVIVS.text()
        VVIVRP = self.VVIVRP.text()
        VVIH = self.VVIH.text()
        VVIRS = self.VVIRS.text()
        try:
            # ranges for the vvi. If range is not met then show invalid error
            if ((((int(VVILRL)) >= 30) and (int(VVILRL)) <= 49) and (int(VVILRL) % 5 == 0)) or ((((int(VVILRL)) >= 50) and (int(VVILRL)) <= 89) and (int(VVILRL) % 1 == 0)) or (((int((VVILRL)) >= 90) and ((int(VVILRL)) <= 175) and (int(VVILRL) % 5 == 0))):
                if(((int(VVIURL)) >= 50) and ((int(VVIURL)) <=175) and (int(VVIURL) % 5 == 0)):
                    if(((float(VVIVA)) >= 0.1) and ((float(VVIVA)) <=5) and (10*(float(VVIVA)) % 1 == 0) or(int(VVIVA)) == 0):
                        if(((int(VVIVPW)) >= 1) and ((int(VVIVPW)) <=30) and (int(VVIVPW) % 1 == 0)) :
                            if((((int(VVIVS) >=0)) and (int(VVIVS) <=5) and (10*float(VVIVS) %1 == 0))) :
                                if((int(VVIVRP) >= 150) and (int(VVIVRP) <= 500) and ((int(VVIVRP)) % 10 == 0)):
                                    if (((((int(VVIH)) >= 30) and ((int(VVIH)) <= 49) and (int(VVIH) % 5 == 0))) or ((((int(VVIH)) >= 50) and ((int( VVIH)) <= 89) and (int(VVIH) % 1 == 0))) or ((((int(VVIH)) >= 90) and ((int(VVIH)) <= 175) and (int(VVIH) % 5 == 0))) or ((int(VVIH) == 0))):
                                        if(((int(VVIRS) >= 0) and ((int(VVIRS) <= 21) and ((int(VVIRS)) % 3 == 0)))):

                                                self.INVALID.setVisible(False)
                                                self.SUCCESS.setVisible(True)
                                                getecgplot_atrial(480,"VVI Plot")
                                                getecgplot_ventricular(480,"VVI Plot")

                                                wb = load_workbook("AccountData.xlsx")

                                                try:
                                                    ws = wb[username]
                                                    ws['E2'] = VVILRL
                                                    ws['E3'] = VVIURL
                                                    ws['E6'] = VVIVA
                                                    ws['E7'] = VVIVPW
                                                    ws['E8'] = VVIVS
                                                    ws['E13'] = VVIVRP
                                                    ws['E11'] =VVIH
                                                    ws['E19'] =VVIRS

                                                    wb.save("AccountData.xlsx")
                                                    print("Added VVI Data")
                                                
                                                except:
                                                    print("Couldn't add the number")


                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)

                                else:
                                    self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                        else:
                             self.INVALID.setVisible(True)
                             self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)

    def transferfunction(self):
        #SET AAI VALUES AND PUSH THEM TO BOARD
        # try:
        frdm_port = "COM3"

        wb = load_workbook("AccountData.xlsx")
        ws = wb[username]

        VVILRL=ws['E2'] 
        VVIURL=ws['E3'] 
        VVIVA=ws['E6'] 
        VVIVPW=ws['E7'] 
        VVIVS=ws['E8'] 
        VVIVRP=ws['E13'] 
        VVIH=ws['E11'] 
        VVIRS=ws['E19'] 

        Start = b'\x16' #putting a value in bits into a variable bcuz the var will be added to the byte stream and sent over to simulink
        SYNC = b'\x22'
        #Fn_set = b'\x55'
        Fn_set = struct.pack("B",55)
        p_aPaceWidth = struct.pack('B', int(VVIVPW.value))
        p_vPaceWidth = struct.pack('B', int(VOOVPW.value))
        p_aPaceAmp = struct.pack('f', float(VVIVA.value))#f is single
        p_vPaceAmp = struct.pack('f', float(AAIAPW.value))
        p_atrialsensitivity = struct.pack('f', float(VVIVS.value))
        p_ventriclesensitivity = struct.pack('f', 9)
        p_ARP = struct.pack('H', int(AAIARP.value)) #H is uint16
        p_VRP = struct.pack('H', int(AAIPVARP.value))
        p_lowratelimit = struct.pack('B', int(VVILRL.value)) #we used B becasue its uint8
        p_Mode = struct.pack('B', 4)
        #blue_en = struct.pack("B", 1)

        Signal_set = Start + Fn_set + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode
        Signal_echo = Start + SYNC + p_aPaceWidth + p_vPaceWidth + p_aPaceAmp + p_vPaceAmp + p_atrialsensitivity + p_ventriclesensitivity + p_ARP + p_VRP + p_lowratelimit + p_Mode

        with serial.Serial(frdm_port, 115200) as pacemaker: #pushes signal to the baord
            pacemaker.write(Signal_set)

        with serial.Serial(frdm_port, 115200) as pacemaker: #send the byte stream that echos to the pacemaker (the block that sends parameter back to you)
            pacemaker.write(Signal_echo)
            data = pacemaker.read(24)
            p_aPaceWidth = data[0]
            p_vPaceWidth = data[1]
            p_aPaceAmp = struct.unpack('f', data[2:6])[0] 
            p_vPaceAmp = struct.unpack('f', data[6:10])[0]
            p_atrialsensitivity = struct.unpack('f', data[10:14])[0]
            p_ventriclesensitivity = struct.unpack('f', data[14:18])[0]
            p_ARP = struct.unpack('H', data[18:20])[0]  #H is uint16
            p_VRP = struct.unpack('H', data[20:22])[0] 
            p_lowratelimit = data[22] #we used B becasue its uint8
            p_Mode = data[23]
            

        print("From the board:")
        print("p_aPaceWidth = ", p_aPaceWidth)
        print("p_vPaceWidth = ", p_vPaceWidth)
        print("p_aPaceAmp = ", p_aPaceAmp)
        print("p_vPaceAmp = ",  p_vPaceAmp)
        print("p_atrialsensitivity = ", p_atrialsensitivity)
        print("p_ventriclesensitivity = ",  p_ventriclesensitivity)
        print("p_ARP = ",  p_ARP)
        print("p_VRP = ",  p_VRP)
        print("p_lowratelimit = ",  p_lowratelimit)
        print("p_Mode = ",  p_Mode)  

    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)


class AOOR(QDialog):
    def __init__(self):
        super(AOOR, self).__init__()
        loadUi("AOOR.ui", self)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.AOORsubmitbutton.clicked.connect(self.AOORinputfunction)
        self.backbutton.clicked.connect(self.backfunction)
        widget.setFixedWidth(900)
        widget.setFixedHeight(700)


    def AOORinputfunction(self):
        AOORLRL = self.AOORLRL.text()
        AOORURL = self.AOORURL.text()
        AOORAA = self.AOORAA.text()
        AOORPW = self.AOORPW.text()
        AOORMSR = self.AOORMSR.text()
        ACTIVEAOOR = self.ACTIVEAOOR.text()
        ReactTimeAOOR = self.ReactTimeAOOR.text()
        RespFactAOOR = self.RespFactAOOR.text()
        RecovTimeAOOR = self.RecovTimeAOOR.text()

        # ranges for the AOOR. If range is not met then show invalid error
        try:
            if ((((int(AOORLRL)) >= 30) and (int(AOORLRL)) <= 49) and (int(AOORLRL) % 5 == 0)) or ((((int(AOORLRL)) >= 50) and (int(AOORLRL)) <= 89) and (int(AOORLRL) % 1 == 0)) or (((int((AOORLRL)) >= 90) and ((int(AOORLRL)) <= 175) and (int(AOORLRL) % 5 == 0))):
                if(((int(AOORURL)) >= 50) and (int(AOORURL)) <=175) and (int(AOORURL) % 5 == 0):
                    if(((((float(AOORAA)) >= 0.1) and (float(AOORAA)) <=5) and (10*(float(AOORAA)) % 1 == 0)) or (float(AOORAA)) == 0) :
                        if((((int(AOORPW)) >= 1) and ((int(AOORPW)) <=30) and (int(AOORPW) % 1 == 0))) :
                            if((int(AOORMSR) >= 50)  and (int(AOORMSR) <= 175) and (int(AOORMSR) % 5 == 0)) :
                                if( ((int(ACTIVEAOOR) >= 1) and ((int(ACTIVEAOOR) <=7) and ((int(ACTIVEAOOR)) % 1 == 0)))):
                                    if ((((int(ReactTimeAOOR)) >= 10) and ((int(ReactTimeAOOR)) <= 50) and (int(ReactTimeAOOR) % 10 == 0))) :
                                        if((((int(RespFactAOOR) >= 1) and ((int(RespFactAOOR) <= 16) and ((int(RespFactAOOR)) % 1 == 0))))):
                                            if((((int(RecovTimeAOOR) >= 2) and ((int(RecovTimeAOOR) <= 16) and ((int(RecovTimeAOOR)) % 1 == 0))))):

                                                self.INVALID.setVisible(False)
                                                self.SUCCESS.setVisible(True)
                                                # open to the file and write the inputed numbers
                                                db = open("Storage/AOOR.txt", "w")
                                                db.write(AOORLRL + "\n" + AOORURL + "\n" + AOORAA + "\n" + AOORPW + "\n" + AOORMSR + "\n" + ACTIVEAOOR + "\n" + ReactTimeAOOR + "\n" + RespFactAOOR + "\n" + RecovTimeAOOR)
                                                print("Success")
                                                db.close()
                                            else:
                                                self.INVALID.setVisible(True)
                                                self.SUCCESS.setVisible(False)
                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)

                                else:
                                    self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                        else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)


    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class VOOR(QDialog):
    def __init__(self):
        super(VOOR, self).__init__()
        loadUi("VOOR.ui", self)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.VOORsubmitbutton.clicked.connect(self.VOORinputfunction)
        self.backbutton.clicked.connect(self.backfunction)
        widget.setFixedWidth(900)
        widget.setFixedHeight(700)

    def VOORinputfunction(self):
        VOORLRL = self.VOORLRL.text()
        VOORURL = self.VOORURL.text()
        VOORVA = self.VOORVA.text()
        VOORPW = self.VOORPW.text()
        VOORMSR = self.VOORMSR.text()
        ACTIVEVOOR = self.ACTIVEVOOR.text()
        ReactTimeVOOR = self.ReactTimeVOOR.text()
        RespFactVOOR = self.RespFactVOOR.text()
        RecovTimeVOOR = self.RecovTimeVOOR.text()

        # ranges for the VOOR. If range is not met then show invalid error
        try:
            if ((((int(VOORLRL)) >= 30) and (int(VOORLRL)) <= 49) and (int(VOORLRL) % 5 == 0)) or ((((int(VOORLRL)) >= 50) and (int(VOORLRL)) <= 89) and (int(VOORLRL) % 1 == 0)) or (((int((VOORLRL)) >= 90) and ((int(VOORLRL)) <= 175) and (int(VOORLRL) % 5 == 0))):
                if(((int(VOORURL)) >= 50) and (int(VOORURL)) <=175) and (int(VOORURL) % 5 == 0):
                    if(((((float(VOORVA)) >= 0.1) and (float(VOORVA)) <=5) and (10*(float(VOORVA)) % 1 == 0))  or (float(VOORVA)) == 0) :
                        if(((float(VOORPW)) >= 1) and ((float(VOORPW)) <=30) and ((float(VOORPW)) % 1 == 0)) :
                            if((int(VOORMSR) >= 50)  and (int(VOORMSR) <= 175) and (int(VOORMSR) % 5 == 0)) :
                                if (int(ACTIVEVOOR) >= 1) and (int(ACTIVEVOOR) <= 7) and (int(ACTIVEVOOR) %1 == 0):
                                    if ((((int(ReactTimeVOOR)) >= 10) and ((int(ReactTimeVOOR)) <= 50) and (int(ReactTimeVOOR) % 10 == 0))) :
                                        if(((int(RespFactVOOR) >= 1) and ((int(RespFactVOOR) <= 16) and ((int(RespFactVOOR) % 1 == 0))))):
                                            if((((int(RecovTimeVOOR) >= 2) and ((int(RecovTimeVOOR) <= 16) and ((int(RecovTimeVOOR)) % 1 == 0))))):

                                                self.INVALID.setVisible(False)
                                                self.SUCCESS.setVisible(True)
                                                    # open to the file and write the inputed numbers
                                                db = open("Storage/VOOR.txt", "w")
                                                db.write(VOORLRL + "\n" +VOORURL + "\n" + VOORVA + "\n" +VOORPW + "\n" + VOORMSR + "\n" + ACTIVEVOOR + "\n" + ReactTimeVOOR + "\n" + RespFactVOOR + "\n" + RecovTimeVOOR)
                                                print("Success")
                                                db.close()

                                            else:
                                                self.INVALID.setVisible(True)
                                                self.SUCCESS.setVisible(False)
                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)

                                else:
                                    self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                        else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)

                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)



    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class AAIR(QDialog):
    def __init__(self):
        super(AAIR, self).__init__()
        loadUi("AAIR.ui", self)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.AAIRsubmitbutton.clicked.connect(self.AAIRinputfunction)
        self.backbutton.clicked.connect(self.backfunction)
        widget.setFixedWidth(1200)
        widget.setFixedHeight(900)

    def AAIRinputfunction(self):
        AAIRLRL = self.AAIRLRL.text()
        AAIRURL = self.AAIRURL.text()
        AAIRAA = self.AAIRAA.text()
        AAIRPW = self.AAIRPW.text()
        AAIRMSR = self.AAIRMSR.text()
        AAIRAS = self.AAIRAS.text()
        AAIRARP = self.AAIRARP.text()
        ReactTimeAAIR = self.ReactTimeAAIR.text()
        AAIRPVARP = self.AAIRPVARP.text()
        AAIRHY = self.AAIRHY.text()
        AAIRRS = self.AAIRRS.text()
        RespFactAAIR = self.RespFactAAIR.text()
        ACTIVEAAIR = self.ACTIVEAAIR.text()
        RecovTimeAAIR = self.RecovTimeAAIR.text()

        # ranges for the AAIR. If range is not met then show invalid error
        try:
            if ((((int(AAIRLRL)) >= 30) and (int(AAIRLRL)) <= 49) and (int(AAIRLRL) % 5 == 0)) or ((((int(AAIRLRL)) >= 50) and (int(AAIRLRL)) <= 89) and (int(AAIRLRL) % 1 == 0)) or (((int((AAIRLRL)) >= 90) and ((int(AAIRLRL)) <= 175) and (int(AAIRLRL) % 5 == 0))):
                if(((int(AAIRURL)) >= 50) and (int(AAIRURL)) <=175) and (int(AAIRURL) % 5 == 0):
                    if(((((float(AAIRAA)) >= 0.1) and (float(AAIRAA)) <=5) and (10*(float(AAIRAA)) % 1 == 0))  or ((float(AAIRAA)) == 0)) :
                        if(((int(AAIRPW)) >= 1) and ((int(AAIRPW)) <=30) and (int(AAIRPW) % 1 == 0)) :
                            if((int(AAIRMSR) >= 50)  and (int(AAIRMSR) <= 175) and (int(AAIRMSR) % 5 == 0)) :
                                if(((float(AAIRAS) >= 0) and (float(AAIRAS) <= 5) and (10*float(AAIRAS) % 1 == 0))):
                                    if(((int(AAIRARP) >= 150) and ((int(AAIRARP) <=500) and ((int(AAIRARP)) % 10 == 0)))):
                                        if ((((int(ReactTimeAAIR)) >= 10) and ((int(ReactTimeAAIR)) <= 50) and (int(ReactTimeAAIR) % 10 == 0))) :
                                            if(((int(AAIRPVARP) >= 150) and ((int(AAIRPVARP) <=500) and ((int(AAIRPVARP)) % 10 == 0)))):
                                                if ((((int(AAIRHY)) >= 30) and (int(AAIRHY)) <= 49) and (int(AAIRHY) % 5 == 0)) or ((((int(AAIRHY)) >= 50) and (int(AAIRHY)) <= 89) and (int(AAIRHY) % 1 == 0)) or (((int((AAIRHY)) >= 90) and ((int(AAIRHY)) <= 175) and (int(AAIRHY) % 5 == 0))):
                                                    if ((((int(AAIRRS)) >= 0) and ((int(AAIRRS)) <= 21) and((int(AAIRRS))%3 ==0))):
                                                        if(((int(RespFactAAIR)) >= 1) and ((int(RespFactAAIR)) <= 16) and((int(RespFactAAIR))% 1 == 0 )):
                                                            if(((int(ACTIVEAAIR)) >= 1) and ((int(ACTIVEAAIR)) <= 7) and((int(ACTIVEAAIR))% 1 == 0 )):
                                                                if((((int(RecovTimeAAIR) >= 2) and ((int(RecovTimeAAIR) <= 16) and ((int(RecovTimeAAIR)) % 1 == 0))))):
                                                                    self.INVALID.setVisible(False)
                                                                    self.SUCCESS.setVisible(True)
                                                                        # open to the file and write the inputed numbers
                                                                    db = open("Storage/AAIR.txt", "w")
                                                                    db.write(AAIRLRL + "\n" + AAIRURL + "\n" +AAIRAA + "\n" + AAIRMSR + "\n" + AAIRPW + "\n"  +  AAIRAS + "\n" + AAIRARP + "\n" + ReactTimeAAIR + "\n" + AAIRPVARP + "\n" + AAIRHY +"\n" + AAIRRS +"\n" + RespFactAAIR + "\n" + ACTIVEAAIR + "\n" + RecovTimeAAIR)
                                                                    print("Success")
                                                                    db.close()
                                                                else:
                                                                    self.INVALID.setVisible(True)
                                                                    self.SUCCESS.setVisible(False)
                                                            else:
                                                                self.INVALID.setVisible(True)
                                                                self.SUCCESS.setVisible(False)
                                                        else:
                                                            self.INVALID.setVisible(True)
                                                            self.SUCCESS.setVisible(False)

                                                    else:
                                                        self.INVALID.setVisible(True)
                                                        self.SUCCESS.setVisible(False)
                                                else:
                                                    self.INVALID.setVisible(True)
                                                    self.SUCCESS.setVisible(False)

                                            else:
                                                self.INVALID.setVisible(True)
                                                self.SUCCESS.setVisible(False)

                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)
                                else:
                                    self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)
                        else:
                            self.INVALID.setVisible(True)
                            self.SUCCESS.setVisible(False)
                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)


    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

class VVIR(QDialog):
    def __init__(self):
        super(VVIR, self).__init__()
        loadUi("VVIR.ui", self)
        self.INVALID.setVisible(False)
        self.SUCCESS.setVisible(False)
        self.VVIRsubmitbutton.clicked.connect(self.VVIRinputfunction)
        self.backbutton.clicked.connect(self.backfunction)
        widget.setFixedWidth(1200)
        widget.setFixedHeight(800)

    def VVIRinputfunction(self):
        VVIRLRL = self.VVIRLRL.text()
        VVIRURL = self.VVIRURL.text()
        VVIRVA = self.VVIRVA.text()
        VVIRPW = self.VVIRPW.text()
        VVIRMSR = self.VVIRMSR.text()
        VVIRVS = self.VVIRVS.text()
        VVIRIVRP = self.VVIRIVRP.text()
        VVIRReactTime = self.VVIRReactTime.text()
        VVIRHY = self.VVIRHY.text()
        VVIRRS = self.VVIRRS.text()
        RespFactVVIR = self.RespFactVVIR.text()
        ACTIVEVVIR = self.ACTIVEVVIR.text()
        RecovTimeVVIR = self.RecovTimeVVIR.text()

        # ranges for the VVIR. If range is not met then show invalid error
        try:
            if ((((int(VVIRLRL)) >= 30) and (int(VVIRLRL)) <= 49) and (int(VVIRLRL) % 5 == 0)) or ((((int(VVIRLRL)) >= 50) and (int(VVIRLRL)) <= 89) and (int(VVIRLRL) % 1 == 0)) or (((int((VVIRLRL)) >= 90) and ((int(VVIRLRL)) <= 175) and (int(VVIRLRL) % 5 == 0))):
                if(((int(VVIRURL)) >= 50) and (int(VVIRURL)) <=175) and (int(VVIRURL) % 5 == 0):
                    if(((float(VVIRVA)) >= 0.1) and ((float(VVIRVA)) <=5) and (10*(float(VVIRVA)) % 1 == 0)):
                        if(((int(VVIRPW)) >= 1) and ((int(VVIRPW)) <=30) and (int(VVIRPW) % 1 == 0)) :
                            if((int(VVIRMSR) >= 50)  and (int(VVIRMSR) <= 175) and (int(VVIRMSR) % 5 == 0)) :
                                if (int(VVIRVS) >= 0) and (int(VVIRVS) <= 5) and (10*float(VVIRVS) % 1 == 0):
                                    if(((int(VVIRIVRP) >= 150) and ((int(VVIRIVRP) <=500) and ((int(VVIRIVRP)) % 10 == 0)))):
                                        if ((((int(VVIRReactTime)) >= 10) and ((int(VVIRReactTime)) <= 50) and (int(VVIRReactTime) % 10 == 0))) :
                                            if(((int(VVIRHY) >= 150) and ((int(VVIRHY) <=500) and ((int(VVIRHY)) % 10 == 0)))):
                                                if ((((int(VVIRRS)) >= 30) and (int(VVIRRS)) <= 49) and (int(VVIRRS) % 5 == 0)) or ((((int(VVIRRS)) >= 50) and (int(VVIRRS)) <= 89) and (int(VVIRRS) % 1 == 0)) or (((int((VVIRRS)) >= 90) and ((int(VVIRRS)) <= 175) and (int(VVIRRS) % 5 == 0))):
                                                    if(((int(RespFactVVIR)) >= 1) and ((int(RespFactVVIR)) <= 16) and((int(RespFactVVIR))% 1 == 0 )):
                                                        if(((int(ACTIVEVVIR)) >= 1) and ((int(ACTIVEVVIR)) <= 7) and((int(ACTIVEVVIR))% 1 == 0 )):
                                                            if((((int(RecovTimeVVIR) >= 2) and ((int(RecovTimeVVIR) <= 16) and ((int(RecovTimeVVIR)) % 1 == 0))))):
                                                                self.INVALID.setVisible(False)
                                                                self.SUCCESS.setVisible(True)
                                                                # open to the file and write the inputed numbers
                                                                db = open("Storage/VVIR.txt", "w")
                                                                db.write(VVIRLRL + "\n" + VVIRURL + "\n" + VVIRVA + "\n" + VVIRPW + "\n" + VVIRMSR + "\n"  + VVIRVS + "\n" + VVIRIVRP+ "\n" + VVIRReactTime + "\n" + VVIRHY + "\n" + VVIRRS +"\n" + RespFactVVIR +"\n" + ACTIVEVVIR + "\n" + RecovTimeVVIR)
                                                                print("Success")
                                                                db.close()
                                                            else:
                                                                self.INVALID.setVisible(True)
                                                                self.SUCCESS.setVisible(False)
                                                        else:
                                                            self.INVALID.setVisible(True)
                                                            self.SUCCESS.setVisible(False)
                                                    else:
                                                        self.INVALID.setVisible(True)
                                                        self.SUCCESS.setVisible(False)

                                                else:
                                                    self.INVALID.setVisible(True)
                                                    self.SUCCESS.setVisible(False)

                                            else:
                                                self.INVALID.setVisible(True)
                                                self.SUCCESS.setVisible(False)

                                        else:
                                            self.INVALID.setVisible(True)
                                            self.SUCCESS.setVisible(False)
                                    else:
                                        self.INVALID.setVisible(True)
                                        self.SUCCESS.setVisible(False)
                                else:
                                    self.INVALID.setVisible(True)
                                    self.SUCCESS.setVisible(False)
                            else:
                                self.INVALID.setVisible(True)
                                self.SUCCESS.setVisible(False)
                        else:
                            self.INVALID.setVisible(True)
                            self.SUCCESS.setVisible(False)
                    else:
                        self.INVALID.setVisible(True)
                        self.SUCCESS.setVisible(False)
                else:
                    self.INVALID.setVisible(True)
                    self.SUCCESS.setVisible(False)
            else:
                self.INVALID.setVisible(True)
                self.SUCCESS.setVisible(False)
        except:
            self.INVALID.setVisible(True)
            self.SUCCESS.setVisible(False)

    def backfunction(self):
        back = Dash()
        widget.addWidget(back)
        widget.setCurrentIndex(widget.currentIndex()+1)

app = QApplication(sys.argv)
mainwindow = Mainscreen()
widget = QtWidgets.QStackedWidget()
widget.addWidget(mainwindow)
widget.show()

app.exec()
